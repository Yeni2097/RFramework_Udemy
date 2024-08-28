from robot.libraries.BuiltIn import BuiltIn
import openpyxl
#ESC + ALT y seleciono los que quiero cambiar para que los nombres se cambien todos al mismo tiempo

archivo=openpyxl.load_workbook(r"D:\Users\yenifer.alfaro\RF_Nivel1\RFramework_Udemy\Documentos\Datos Inicio Sesion PY.xlsx")

def Numero_Filas(hoja):
    print("Esta entrando al num de filas****")
    HojaActiva = archivo[hoja]
    return  HojaActiva.max_row

def Dato_por_columna(hoja, fila, col):
    hojaActiva = archivo[hoja]
    valor = hojaActiva.cell(int(fila), int(col)).value
    BuiltIn().log_to_console("El valor es****"+str(valor))
    return valor

#print(archivo.sheetnames)
#print("pagina activa: " + archivo.active.title)
#HojaActiva=archivo["Hoja1"]
#print(ac['A2'].value)
#print(ac['B2'].value)
#print(ac['C2'].value)

#filas = HojaActiva.max_row
#columnas = HojaActiva.max_column

#for reglon in HojaActiva['A2':'B7']:
    #for columna in reglon:
        #print(columna.value)

#for f in HojaActiva[1:filas+1]:
    #for c in f:
        #print(c.value)
